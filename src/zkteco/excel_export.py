"""Excel report generator for the ZKTeco time tracker.

Produces a formatted ``.xlsx`` workbook with:

- a **Resumen** sheet: one row per user for the current week (hours worked,
  remaining to reach the 30h goal) plus a global count of completed weeks;
- a **General** sheet: a week-by-user matrix colored by whether the 30h goal
  was met, with totals per week and per user (date-filterable);
- one **sheet per user** (tab titled ``Nombre (ID)``) with their weekly
  history, every session in the filtered range (date, in, out, duration) and a
  bar chart of weekly hours against the 30h goal.

The date range passed to :func:`build_workbook` filters the sessions used for
the general matrix and the per-user sheets; the Resumen sheet always reflects
the *current* week.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .report import (
    WEEKLY_REQUIREMENT_SECONDS,
    week_start_for,
)

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
SECTION_FILL = PatternFill("solid", fgColor="DBEAFE")
GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
GREY_FILL = PatternFill("solid", fgColor="F1F5F9")
BLUE_FILL = PatternFill("solid", fgColor="DBEAFE")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1E3A8A")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

MAX_SHEET = 31  # Excel worksheet name limit


def _hours(seconds: float) -> float:
    return seconds / 3600.0


def _clean_sheet_name(name: str, fallback: str) -> str:
    cleaned = "".join(c for c in name if c not in r"[]:*?/\\").strip()
    if not cleaned:
        cleaned = fallback
    return cleaned[:MAX_SHEET]


def _style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _style_section(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BOLD
        cell.fill = SECTION_FILL
        cell.border = BORDER


def _set_widths(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _bar_chart(ws, rows: int, data_col: int, ref_col: int, title: str, categories_col: int = 1) -> BarChart:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Horas"
    chart.x_axis.title = "Semana"
    data = Reference(ws, min_col=data_col, min_row=1, max_col=ref_col, max_row=rows)
    cats = Reference(ws, min_col=categories_col, min_row=2, max_row=rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 10
    return chart


def _user_session_rows(db, user_id: str, start: date, end: date) -> list:
    rows = []
    for s in db.sessions_for(user_id=user_id):
        day = datetime.fromisoformat(s["clock_in_at"]).date()
        if start <= day <= end:
            rows.append(s)
    return rows


def build_workbook(db, start: date, end: date) -> Workbook:
    wb = Workbook()

    weeks = []
    wk = week_start_for(start)
    end_wk = week_start_for(end)
    while wk <= end_wk:
        weeks.append(wk)
        wk += timedelta(days=7)

    users = [u for u in db.all_users() if u["active"]]
    req = db.weekly_requirement

    # weekly totals per user within the filter (inclusive), plus running session
    totals: dict[tuple[str, date], float] = defaultdict(float)
    for s in db.sessions_in_range(start, end):
        day = datetime.fromisoformat(s["clock_in_at"]).date()
        w = week_start_for(day)
        totals[(s["user_id"], w)] += float(s["session_seconds"] or 0)
    for u in users:
        open_ = db.open_session(u["user_id"])
        if open_:
            open_start = datetime.fromisoformat(open_["clock_in_at"])
            if start <= open_start.date() <= end:
                w = week_start_for(open_start.date())
                totals[(u["user_id"], w)] += max(
                    0.0, (datetime.now() - open_start).total_seconds()
                )

    _build_resumen_sheet(wb, db, users, req)
    _build_general_sheet(wb, db, users, req, weeks, totals)
    for u in users:
        _build_user_sheet(wb, db, u, req, weeks, totals, start, end)

    return wb


def _build_resumen_sheet(wb: Workbook, db, users, req: float) -> None:
    ws = wb.active
    ws.title = "Resumen"
    _set_widths(ws, {1: 28, 2: 12, 3: 14, 4: 14, 5: 16, 6: 20})

    now = datetime.now()
    current = week_start_for(now)

    ws["A1"] = "Reporte de asistencia — ZKTeco"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Semana en curso: {current.day:02d}/{current.month:02d} al "
        f"{(current + timedelta(days=6)).day:02d}/{(current + timedelta(days=6)).month:02d} "
        f"· generado {now:%d/%m/%Y %H:%M}"
    )
    ws["A2"].font = Font(color="FF64748B")

    row = 4
    ws.cell(row=row, column=1, value="Estado actual de la semana (meta 30h)").font = BOLD
    row += 1
    headers = ["Nombre", "ID", "Horas semana", "Meta (h)", "Faltan (h)", "Estado"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(headers))
    header_row = row
    row += 1

    completed = 0
    for u in users:
        uid = u["user_id"]
        secs = _current_week_seconds(db, uid, current)
        hours = _hours(secs)
        remaining = max(0.0, _hours(req) - hours)
        state = "Completada" if secs >= req else (
            "En progreso" if secs > 0 else "Sin horas"
        )
        fill = GREEN_FILL if secs >= req else (RED_FILL if secs > 0 else GREY_FILL)
        ws.cell(row=row, column=1, value=u["name"] or uid)
        ws.cell(row=row, column=2, value=uid)
        ws.cell(row=row, column=3, value=round(hours, 2))
        ws.cell(row=row, column=4, value=_hours(req))
        ws.cell(row=row, column=5, value=round(remaining, 2))
        ws.cell(row=row, column=6, value=state)
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.fill = fill
            if col >= 3:
                cell.number_format = "0.00"
                cell.alignment = CENTER
        ws.cell(row=row, column=3).number_format = "0.00"
        if secs >= req:
            completed += 1
        row += 1

    ws.cell(row=row, column=1, value=f"Total usuarios con la meta completada: {completed} de {len(users)}").font = BOLD
    row += 2

    # Global completed-weeks summary within the filter
    ws.cell(row=row, column=1, value="Semanas completadas (rango del filtro)").font = BOLD
    row += 1
    sub = ["Usuario", "Semanas completadas", "Semanas con horas", "Total semanas"]
    for col, h in enumerate(sub, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(sub))
    row += 1
    grand_total = 0
    weeks_bounds = _weeks_between_bounds(db, users)
    for u in users:
        uid = u["user_id"]
        met = sum(1 for w in weeks_bounds if _weeks_total(db, uid, w) >= req)
        with_hours = sum(1 for w in weeks_bounds if _weeks_total(db, uid, w) > 0)
        ws.cell(row=row, column=1, value=u["name"] or uid)
        ws.cell(row=row, column=2, value=met)
        ws.cell(row=row, column=3, value=with_hours)
        ws.cell(row=row, column=4, value=len(weeks_bounds))
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER
            if col >= 2:
                ws.cell(row=row, column=col).alignment = CENTER
        if met:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = GREEN_FILL
        grand_total += met
        row += 1
    ws.cell(row=row, column=1, value=f"Total de semanas completadas (todos los usuarios): {grand_total}").font = BOLD
    ws.freeze_panes = f"A{header_row + 1}"


def _weeks_between_bounds(db, users) -> list[date]:
    """Mondays from the earliest user session to today (inclusive)."""
    first = min((datetime.fromisoformat(s["clock_in_at"]).date()
                 for u in users
                 for s in db.sessions_for(user_id=u["user_id"])), default=None)
    if first is None:
        return []
    weeks: list[date] = []
    wk = week_start_for(first)
    end = week_start_for(datetime.now())
    while wk <= end:
        weeks.append(wk)
        wk += timedelta(days=7)
    return weeks


def _weeks_total(db, user_id: str, wk: date) -> float:
    """Total seconds for ``user_id`` in the week starting ``wk``, including a
    running session if any."""
    total = 0.0
    for s in db.sessions_for(user_id=user_id):
        day = datetime.fromisoformat(s["clock_in_at"]).date()
        if week_start_for(day) == wk:
            total += float(s["session_seconds"] or 0)
    open_ = db.open_session(user_id)
    if open_ and week_start_for(datetime.fromisoformat(open_["clock_in_at"]).date()) == wk:
        total += max(
            0.0, (datetime.now() - datetime.fromisoformat(open_["clock_in_at"])).total_seconds()
        )
    return total


def _current_week_seconds(db, user_id: str, current: date) -> float:
    return _weeks_total(db, user_id, current)


def _build_general_sheet(
    wb: Workbook, db, users, req: float, weeks: list[date], totals: dict
) -> None:
    ws = wb.create_sheet("General")
    _set_widths(ws, {1: 26})
    for i in range(len(weeks)):
        _set_widths(ws, {i + 2: 14})

    ws["A1"] = "Matriz de semanas completadas por usuario"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Del {weeks[0].day:02d}/{weeks[0].month:02d}/{weeks[0].year} al "
        f"{(weeks[-1] + timedelta(days=6)).day:02d}/{(weeks[-1] + timedelta(days=6)).month:02d}/{weeks[-1].year}"
        if weeks else "Sin semanas en el rango"
    )
    ws["A2"].font = Font(color="FF64748B")

    current = week_start_for(datetime.now())
    row = 4
    ws.cell(row=row, column=1, value="Usuario")
    for col, w in enumerate(weeks, start=2):
        ws.cell(row=row, column=col, value=w.strftime("%d/%m"))
    ws.cell(row=row, column=len(weeks) + 2, value="Completadas")
    _style_header(ws, row, len(weeks) + 2)
    header_row = row
    row += 1

    per_user_completed: list[int] = []
    for u in users:
        uid = u["user_id"]
        ws.cell(row=row, column=1, value=u["name"] or uid)
        count = 0
        for col, w in enumerate(weeks, start=2):
            secs = totals.get((uid, w), 0.0)
            cell = ws.cell(row=row, column=col, value=round(_hours(secs), 2) if secs else "")
            cell.number_format = "0.00"
            cell.border = BORDER
            if secs >= req:
                cell.fill = GREEN_FILL
                count += 1
            elif secs > 0:
                cell.fill = RED_FILL
            elif w == current:
                cell.fill = BLUE_FILL
            else:
                cell.fill = GREY_FILL
            cell.alignment = CENTER
        comp = ws.cell(row=row, column=len(weeks) + 2, value=count)
        comp.border = BORDER
        comp.fill = GREEN_FILL if count else GREY_FILL
        comp.alignment = CENTER
        per_user_completed.append(count)
        row += 1

    ws.cell(row=row, column=1, value="Total usuarios cumpliendo")
    for col, w in enumerate(weeks, start=2):
        cell = ws.cell(
            row=row, column=col,
            value=sum(1 for u in users if totals.get((u["user_id"], w), 0.0) >= req),
        )
        cell.border = BORDER
        cell.alignment = CENTER
        cell.fill = SECTION_FILL
    ws.cell(row=row, column=len(weeks) + 2, value=sum(per_user_completed)).border = BORDER

    ws.freeze_panes = f"B{header_row + 1}"


def _build_user_sheet(
    wb: Workbook, db, u, req: float, weeks: list[date], totals: dict, start: date, end: date
) -> None:
    uid = u["user_id"]
    name = _clean_sheet_name(f"{u['name'] or uid} ({uid})", uid)
    ws = wb.create_sheet(name)
    _set_widths(ws, {1: 34, 2: 14, 3: 14, 4: 14})

    ws["A1"] = f"{u['name'] or uid}  ·  ID {uid}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Del {start.day:02d}/{start.month:02d}/{start.year} al {end.day:02d}/{end.month:02d}/{end.year}"
    ws["A2"].font = Font(color="FF64748B")

    row = 4
    ws.cell(row=row, column=1, value="Historial semanal").font = BOLD
    row += 1
    headers = ["Semana", "Horas", "Meta (h)", "Estado"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(headers))
    row += 1
    first_weekly_row = row
    for w in weeks:
        secs = totals.get((uid, w), 0.0)
        end_date = w + timedelta(days=6)
        ws.cell(row=row, column=1, value=f"{w.day:02d}/{w.month:02d} – {end_date.day:02d}/{end_date.month:02d}")
        ws.cell(row=row, column=2, value=round(_hours(secs), 2))
        ws.cell(row=row, column=3, value=_hours(req))
        ws.cell(row=row, column=4, value="Completada" if secs >= req else ("En curso" if secs > 0 else "Sin horas"))
        fill = GREEN_FILL if secs >= req else (RED_FILL if secs > 0 else GREY_FILL)
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.fill = fill
        ws.cell(row=row, column=2).number_format = "0.00"
        ws.cell(row=row, column=3).number_format = "0.00"
        row += 1
    last_weekly_row = row - 1

    if last_weekly_row >= first_weekly_row:
        chart = _bar_chart(
            ws,
            last_weekly_row - first_weekly_row + 2,
            data_col=2,
            ref_col=3,
            title="Horas por semana vs meta 30h",
        )
        ws.add_chart(chart, f"F{first_weekly_row}")

    row += 1
    ws.cell(row=row, column=1, value="Sesiones del período").font = BOLD
    row += 1
    s_headers = ["Fecha", "Entrada", "Salida", "Duración"]
    for col, h in enumerate(s_headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(s_headers))
    row += 1
    sessions = _user_session_rows(db, uid, start, end)
    for s in sessions:
        clock_in = datetime.fromisoformat(s["clock_in_at"])
        out = datetime.fromisoformat(s["clock_out_at"]) if s["clock_out_at"] else None
        ws.cell(row=row, column=1, value=clock_in.date())
        ws.cell(row=row, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=2, value=clock_in.time().strftime("%H:%M:%S"))
        ws.cell(row=row, column=3, value=out.time().strftime("%H:%M:%S") if out else "En curso")
        ws.cell(row=row, column=4, value=_format_duration(s["session_seconds"] or 0))
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    if not sessions:
        ws.cell(row=row, column=1, value="Sin sesiones en el período.").font = Font(color="FF94A3B8")
    ws.freeze_panes = "A2"


def _format_duration(seconds: float) -> str:
    seconds = int(seconds)
    hours, rem = divmod(seconds, 3600)
    minutes, secs = divmod(rem, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"
